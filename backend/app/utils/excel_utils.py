from typing import Any, List
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


def apply_border(cell):
    thin = Side(style="thin", color="000000")
    cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)


def style_cell(cell, fill=None, font=None, alignment=None):
    if fill:
        cell.fill = fill
    if font:
        cell.font = font
    if alignment:
        cell.alignment = alignment
    apply_border(cell)


def auto_width(ws):
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            value = "" if cell.value is None else str(cell.value)
            max_len = max(max_len, len(value))
        ws.column_dimensions[col_letter].width = min(max_len + 3, 28)


def write_standard_sheet(ws, title: str, headers: List[str], rows: List[List[Any]]):
    dark_blue = PatternFill("solid", fgColor="203864")
    title_fill = PatternFill("solid", fgColor="1F3C73")
    white_font = Font(color="FFFFFF", bold=True)
    header_font = Font(color="FFFFFF", bold=True, size=10)
    body_font = Font(color="000000", size=10)
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)

    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(headers))
    ws.cell(1, 1, title)
    style_cell(ws.cell(1, 1), fill=title_fill, font=white_font, alignment=center)

    for col_idx, header in enumerate(headers, start=1):
        c = ws.cell(2, col_idx, header)
        style_cell(c, fill=dark_blue, font=header_font, alignment=center)

    for r_idx, row in enumerate(rows, start=3):
        for c_idx, value in enumerate(row, start=1):
            c = ws.cell(r_idx, c_idx, value)
            style_cell(c, font=body_font, alignment=center)

    ws.freeze_panes = "A3"
    auto_width(ws)
