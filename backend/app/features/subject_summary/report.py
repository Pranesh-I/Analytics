from __future__ import annotations

from typing import Any, Dict, List

from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side


HEADERS = [
    "Rank", "Roll", "Name", "Subject", "Attempted", "Correct", "Wrong",
    "Unattempted", "Marks/100", "Accuracy%", "Neg.Marks", "Risk Level", "Remarks"
]

SUBJECT_FILL = {
    "Physics": "BDD7EE",
    "Chemistry": "E2F0D9",
    "Mathematics": "EADCF5",
}

RISK_FILL = {
    "Low": "E2F0D9",
    "Moderate": "FFF2CC",
    "High": "F4CCCC",
    "Critical": "EA9999",
}

TITLE_FILL = "1F497D"
SUBTITLE_FILL = "2F75B5"
HEADER_FILL = "2F75B5"

WHITE_FONT = Font(color="FFFFFF", bold=True)
TITLE_FONT = Font(color="FFFFFF", bold=True, size=15)
SUBTITLE_FONT = Font(color="FFFFFF", bold=True, size=11)
thin_side = Side(style="thin", color="000000")
BORDER = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)


def _apply(cell, fill=None, font=None, align=None):
    cell.border = BORDER
    if fill is not None:
        cell.fill = fill
    if font is not None:
        cell.font = font
    if align is not None:
        cell.alignment = align


def write_subject_summary_sheet(ws: Worksheet, subject_rows: List[Dict[str, Any]]) -> None:
    ws.title = "2_Subjectwise"

    ws.merge_cells("A1:M1")
    ws.merge_cells("A2:M2")
    ws["A1"] = "Subject-wise Performance"
    ws["A2"] = "Physics | Chemistry | Mathematics"

    ws["A1"].fill = PatternFill("solid", fgColor=TITLE_FILL)
    ws["A1"].font = TITLE_FONT
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")

    ws["A2"].fill = PatternFill("solid", fgColor=SUBTITLE_FILL)
    ws["A2"].font = SUBTITLE_FONT
    ws["A2"].alignment = Alignment(horizontal="center", vertical="center")

    for col_idx, header in enumerate(HEADERS, start=1):
        cell = ws.cell(row=3, column=col_idx, value=header)
        _apply(
            cell,
            fill=PatternFill("solid", fgColor=HEADER_FILL),
            font=WHITE_FONT,
            align=Alignment(horizontal="center", vertical="center", wrap_text=True),
        )

    start_row = 4
    for row_idx, row in enumerate(subject_rows, start=start_row):
        subject = str(row.get("subject", "")).strip()
        subject_fill = PatternFill("solid", fgColor=SUBJECT_FILL.get(subject, "FFFFFF"))

        values = [
            row.get("rank", ""),
            row.get("roll_no", ""),
            row.get("name", ""),
            row.get("subject", ""),
            row.get("attempted", ""),
            row.get("correct", ""),
            row.get("wrong", ""),
            row.get("unattempted", ""),
            row.get("marks", ""),
            f'{row.get("accuracy", 0):.1f}%',
            row.get("negative_marks", ""),
            row.get("risk_level", ""),
            row.get("remarks", ""),
        ]

        for col_idx, value in enumerate(values, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)

            fill = subject_fill
            if col_idx == 12:
                fill = PatternFill("solid", fgColor=RISK_FILL.get(str(row.get("risk_level", "")), "FFFFFF"))

            _apply(
                cell,
                fill=fill,
                font=Font(color="000000"),
                align=Alignment(horizontal="center", vertical="center", wrap_text=True),
            )

    widths = {
        "A": 8,
        "B": 10,
        "C": 24,
        "D": 16,
        "E": 10,
        "F": 10,
        "G": 10,
        "H": 12,
        "I": 12,
        "J": 11,
        "K": 11,
        "L": 12,
        "M": 18,
    }
    for col, width in widths.items():
        ws.column_dimensions[col].width = width

    ws.freeze_panes = "A4"
    ws.auto_filter.ref = f"A3:M{max(3, 3 + len(subject_rows))}"
    ws.row_dimensions[1].height = 24
    ws.row_dimensions[2].height = 20
    ws.row_dimensions[3].height = 28